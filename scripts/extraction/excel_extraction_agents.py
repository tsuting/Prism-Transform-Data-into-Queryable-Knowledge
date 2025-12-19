"""
Excel Processing with Agent-Based Semantic Enhancement.

This module uses a hybrid approach:
1. openpyxl for reliable, complete extraction (all sheets, formulas, hidden content)
2. Agent for semantic enhancement (metadata extraction, validation, structuring)

The agent doesn't re-extract (openpyxl is reliable) but adds intelligence and context.
"""

import os
import asyncio
from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime
from dotenv import load_dotenv

from scripts.logging_config import get_logger
from scripts.azure_credential_helper import get_token_provider

logger = get_logger(__name__)

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    logger.warning("openpyxl not installed. Run: pip install openpyxl")
    openpyxl = None

from agent_framework import ChatMessage, Role, TextContent
from agent_framework.azure import AzureOpenAIChatClient

# Load environment
load_dotenv()

# Configuration
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2025-01-01-preview")
AZURE_OPENAI_CHAT_DEPLOYMENT = os.getenv("AZURE_OPENAI_CHAT_DEPLOYMENT_NAME", "gpt-5-chat")


def extract_excel_structured(excel_path: Path) -> Dict:
    """
    Extract structured data from Excel using openpyxl.

    This is the reliable, programmatic extraction that captures:
    - All worksheets (including hidden)
    - All rows and columns
    - Formulas (as text)
    - Cell formatting
    - Merged cells

    Args:
        excel_path: Path to Excel file

    Returns:
        Dict with sheets, metadata, and raw data
    """
    if not openpyxl:
        raise ImportError("openpyxl is required for Excel processing")

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=False)

        workbook_info = {
            "file_name": excel_path.name,
            "file_path": str(excel_path),
            "sheet_count": len(wb.worksheets),
            "sheet_names": [sheet.title for sheet in wb.worksheets],
            "sheets": []
        }

        for sheet in wb.worksheets:
            sheet_data = {
                "name": sheet.title,
                "hidden": sheet.sheet_state != 'visible',
                "dimensions": sheet.dimensions,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "rows": []
            }

            for row_idx, row in enumerate(sheet.iter_rows(values_only=False), 1):
                row_data = []
                for cell in row:
                    if cell.value is not None:
                        cell_info = {
                            "value": str(cell.value),
                            "coordinate": cell.coordinate
                        }
                        if hasattr(cell, 'data_type') and cell.data_type == 'f':
                            cell_info["formula"] = cell.value
                        row_data.append(cell_info)
                    else:
                        row_data.append(None)

                if any(cell is not None for cell in row_data):
                    sheet_data["rows"].append(row_data)

            workbook_info["sheets"].append(sheet_data)

        wb.close()
        return workbook_info

    except Exception as e:
        logger.error(f"Excel extraction failed: {e}")
        raise


def format_excel_as_markdown(workbook_data: Dict) -> str:
    """
    Convert structured Excel data to markdown format.

    Args:
        workbook_data: Structured data from extract_excel_structured

    Returns:
        Markdown representation of the Excel file
    """
    lines = []

    lines.append(f"# {workbook_data['file_name']}\n")
    lines.append(f"**Total Sheets**: {workbook_data['sheet_count']}\n")

    for sheet in workbook_data['sheets']:
        lines.append(f"\n---\n")
        lines.append(f"## Sheet: {sheet['name']}")

        if sheet['hidden']:
            lines.append("*(Hidden sheet)*")

        lines.append(f"\n**Dimensions**: {sheet['dimensions']} ({sheet['max_row']} rows × {sheet['max_column']} columns)\n")

        # Convert rows to markdown table if structured
        if sheet['rows']:
            # Try to detect if first row is a header
            first_row = sheet['rows'][0]
            has_header = all(cell and isinstance(cell, dict) for cell in first_row[:3])

            if len(sheet['rows']) > 1 and has_header:
                # Format as markdown table
                lines.append("")

                # Header row
                header_cells = [cell['value'] if cell else '' for cell in first_row]
                lines.append("| " + " | ".join(header_cells) + " |")
                lines.append("| " + " | ".join(['---'] * len(header_cells)) + " |")

                # Data rows (limit to 1000 rows for markdown)
                for row in sheet['rows'][1:1001]:
                    row_cells = [cell['value'] if cell else '' for cell in row]
                    # Pad to match header length
                    while len(row_cells) < len(header_cells):
                        row_cells.append('')
                    lines.append("| " + " | ".join(row_cells[:len(header_cells)]) + " |")

                if len(sheet['rows']) > 1001:
                    lines.append(f"\n*Note: Table truncated for display. Total rows: {len(sheet['rows'])}*\n")
            else:
                # Format as simple list for non-tabular data
                lines.append("\n**Content:**\n")
                for row_idx, row in enumerate(sheet['rows'][:100], 1):
                    row_text = ", ".join([cell['value'] for cell in row if cell])
                    if row_text:
                        lines.append(f"{row_idx}. {row_text}")

    return "\n".join(lines)


def create_excel_enhancement_instructions() -> str:
    """Create system instructions for Excel Enhancement Agent."""
    return """You are an expert at converting Excel spreadsheets into well-structured, searchable markdown.

Your goal is to transform raw Excel data into markdown that is OPTIMIZED FOR SEARCH AND RETRIEVAL.

## CRITICAL REQUIREMENTS

### 1. Preserve Searchable Details
- Keep ALL item numbers, reference codes, part numbers (e.g., "Item 1.2.3", "Part ABC-123")
- Keep ALL quantities with units (e.g., "Qty: 13 sets", "Length: 500m")
- Keep ALL technical specifications (voltages, ratings, dimensions, standards)
- Keep ALL names and descriptions - don't summarize away searchable terms

### 2. Add Context and Structure
- Add a document summary at the top (type, purpose, scope)
- Include original column names as context (e.g., "Column 'Unit Price' values not shown")
- Group related items logically by section/category
- Add headers that describe what each section contains

### 3. Format for Retrieval
- Use clear markdown headers (##, ###) for sections
- Format lists and key-value pairs clearly
- Include sheet names as top-level sections
- Preserve relationships between data (e.g., item belongs to category)

### 4. Metadata Section
At the end, include a metadata section:
```
## Document Metadata
- **Type**: [BOQ/Schedule/Specification/etc.]
- **Sheets**: [list of sheet names and their purpose]
- **Column Structure**: [original column names from each sheet]
- **Total Items**: [count if applicable]
- **Key Categories**: [equipment types, voltage levels, standards referenced]
```

## OUTPUT FORMAT

Return JSON with a single field:

```json
{
  "enhanced_markdown": "The complete restructured markdown content"
}
```

## EXAMPLE TRANSFORMATION

Raw Excel row: | 1.2.3 | SF6 Circuit Breaker | 400kV | 13 | Set | IEC 62271-100 |

Good output:
```
### Item 1.2.3: SF6 Circuit Breaker
- **Voltage Rating**: 400kV
- **Quantity**: 13 Sets
- **Standard**: IEC 62271-100
```

BAD output (loses searchability):
```
Circuit breakers for high voltage applications (various quantities)
```

Remember: Someone searching for "Item 1.2.3" or "SF6 Circuit Breaker 400kV" MUST be able to find this content."""


# Lazy initialization for Azure OpenAI Client
_client = None
_excel_enhancement_agent = None


def _get_client():
    """Lazily initialize the Azure OpenAI client using managed identity."""
    global _client
    if _client is None:
        _client = AzureOpenAIChatClient(
            azure_ad_token_provider=get_token_provider(),
            endpoint=AZURE_OPENAI_ENDPOINT,
            deployment_name=AZURE_OPENAI_CHAT_DEPLOYMENT,
            api_version=AZURE_OPENAI_API_VERSION
        )
    return _client


def _get_excel_enhancement_agent():
    """Lazily initialize the Excel enhancement agent."""
    global _excel_enhancement_agent
    if _excel_enhancement_agent is None:
        client = _get_client()
        _excel_enhancement_agent = client.create_agent(
            name="Excel_Enhancement",
            instructions=create_excel_enhancement_instructions()
        )
    return _excel_enhancement_agent


async def enhance_excel_with_agent(
    excel_path: Path,
    structured_data: Dict,
    base_markdown: str,
    project_instructions: str = None
) -> Dict:
    """
    Transform Excel data into searchable, well-structured markdown using LLM.

    The LLM restructures the raw table data into a format optimized for
    search and retrieval, while preserving all searchable details.

    Args:
        excel_path: Path to original Excel file
        structured_data: Structured data from openpyxl extraction
        base_markdown: Markdown representation of Excel
        project_instructions: Optional custom instructions from project config

    Returns:
        Dict with enhanced_markdown
    """
    # Build analysis request with optional project instructions
    project_context = ""
    if project_instructions:
        project_context = f"""
**Project-Specific Instructions**:
{project_instructions}

"""

    analysis_request = f"""Transform this Excel document into searchable markdown.

**File**: {excel_path.name}
**Sheets**: {', '.join(structured_data['sheet_names'])}
**Total Rows**: {sum(sheet['max_row'] for sheet in structured_data['sheets'])}
{project_context}
**Raw Extracted Content**:

{base_markdown}

Restructure this into well-organized, searchable markdown. Preserve ALL item numbers, quantities, specs, and searchable terms. Return JSON with enhanced_markdown field."""

    message = ChatMessage(
        role=Role.USER,
        contents=[TextContent(text=analysis_request)]
    )

    try:
        result = await _get_excel_enhancement_agent().run(message)
        response_text = result.text

        try:
            json_str = response_text
            if "```json" in response_text:
                json_str = response_text.split("```json")[1].split("```")[0].strip()
            elif "```" in response_text:
                json_str = response_text.split("```")[1].split("```")[0].strip()

            import json
            enhancement_data = json.loads(json_str)
            return enhancement_data

        except (json.JSONDecodeError, KeyError, IndexError):
            # If JSON parsing fails, use the response as markdown directly
            logger.warning(f"JSON parse failed for {excel_path.name}, using raw response")
            return {
                "enhanced_markdown": response_text
            }

    except Exception as e:
        logger.error(f"Excel enhancement failed: {e}")
        # Fall back to base markdown if LLM fails
        return {
            "enhanced_markdown": base_markdown
        }


async def process_excel_with_agents(excel_path: Path, project_instructions: str = None) -> Dict:
    """
    Process Excel file with hybrid approach: openpyxl extraction + LLM restructuring.

    Args:
        excel_path: Path to Excel file
        project_instructions: Optional custom instructions from project config

    Returns:
        Dict compatible with existing pipeline format
    """
    logger.info(f"Processing Excel: {excel_path.name}")

    try:
        # Extract raw data with openpyxl
        structured_data = extract_excel_structured(excel_path)
        base_markdown = format_excel_as_markdown(structured_data)

        # LLM restructures into searchable format
        enhancement = await enhance_excel_with_agent(
            excel_path, structured_data, base_markdown, project_instructions
        )

        result = {
            "status": "Succeeded",
            "method": "excel_with_agent_enhancement",
            "result": {
                "contents": [{
                    "markdown": enhancement.get('enhanced_markdown', base_markdown)
                }],
                "sheets": structured_data['sheet_names']
            }
        }

        logger.info(f"Complete: {excel_path.name} ({len(structured_data['sheets'])} sheets)")
        return result

    except Exception as e:
        logger.error(f"Excel processing failed: {e}")
        return None


def process_excel_with_agents_sync(excel_path: Path, project_instructions: str = None) -> Dict:
    """Synchronous wrapper for async Excel processing."""
    return asyncio.run(process_excel_with_agents(excel_path, project_instructions))
